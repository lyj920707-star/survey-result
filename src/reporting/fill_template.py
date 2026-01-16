"""
템플릿에 분석 결과를 입력하는 스크립트

- 설문 결과 JSON을 읽어 템플릿 J열에 평균값 삽입
- 문항 내용 기반 자동 매핑
- 검토용 리포트 생성
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook
from difflib import SequenceMatcher


def normalize_text(text: str) -> str:
    """텍스트를 정규화하여 비교 가능하게 만듭니다."""
    if not text:
        return ""
    # 공백, 특수문자 제거, 소문자 변환
    text = re.sub(r'[\s\.\,\(\)\[\]\-\_\:\/]', '', str(text))
    return text.lower()


def similarity_ratio(text1: str, text2: str) -> float:
    """두 텍스트의 유사도를 계산합니다."""
    norm1 = normalize_text(text1)
    norm2 = normalize_text(text2)
    return SequenceMatcher(None, norm1, norm2).ratio()


def find_best_match(question: str, template_questions: dict, threshold: float = 0.5) -> tuple:
    """
    질문에 가장 잘 매칭되는 템플릿 행을 찾습니다.

    Returns:
        (row_number, similarity_score, template_question)
    """
    best_match = None
    best_score = 0
    best_question = ""

    for row, template_q in template_questions.items():
        score = similarity_ratio(question, template_q)
        if score > best_score and score >= threshold:
            best_score = score
            best_match = row
            best_question = template_q

    return best_match, best_score, best_question


def extract_template_questions(ws, j_column_rows: list) -> dict:
    """템플릿에서 J열에 값을 입력해야 하는 행들의 질문(H열)을 추출합니다."""
    questions = {}
    for row in j_column_rows:
        h_value = ws[f'H{row}'].value
        if h_value:
            questions[row] = str(h_value)
    return questions


def fill_template(template_path: Path, results: dict, output_path: Path) -> dict:
    """
    템플릿 파일에 분석 결과를 입력합니다.

    Returns:
        매핑 결과 딕셔너리 (검토용)
    """
    wb = load_workbook(template_path)
    ws = wb.active

    # J열에 값을 입력해야 하는 행 목록 (템플릿 구조 기반)
    # Part 2: 정량 평가 (J14~J21, J17 제외 - 주관식)
    # Part 3: 과목별 평가 (J26~J37)
    j_column_rows = [14, 15, 16, 18, 19, 20, 21, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37]

    # 템플릿에서 질문 추출
    template_questions = extract_template_questions(ws, j_column_rows)

    # 매핑 결과 저장 (검토용)
    mapping_results = {
        'matched': [],
        'unmatched_survey': [],
        'unmatched_template': list(j_column_rows)
    }

    # 각 설문 문항에 대해 매핑 시도
    for q_stats in results['questions']:
        question = q_stats['question']
        mean_value = q_stats['mean']

        row, score, template_q = find_best_match(question, template_questions)

        if row:
            # 매핑 성공 - J열에 값 입력
            ws[f'J{row}'] = mean_value
            mapping_results['matched'].append({
                'row': row,
                'survey_question': question[:80],
                'template_question': template_q[:80],
                'similarity': round(score, 2),
                'mean': mean_value
            })
            mapping_results['unmatched_template'].remove(row)
        else:
            # 매핑 실패
            mapping_results['unmatched_survey'].append({
                'survey_question': question,
                'mean': mean_value
            })

    # 출력 디렉토리 생성
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 파일 저장
    wb.save(output_path)

    return mapping_results


def print_verification_report(mapping_results: dict, output_path: Path):
    """매핑 검토 리포트를 출력합니다."""
    print("\n" + "=" * 70)
    print("검토 리포트")
    print("=" * 70)

    # 매핑 성공 항목
    print(f"\n[매핑 성공] {len(mapping_results['matched'])}개 문항")
    print("-" * 70)
    for m in mapping_results['matched']:
        print(f"  J{m['row']:2d} | 평균: {m['mean']:.2f} | 유사도: {m['similarity']:.0%}")
        print(f"       설문: {m['survey_question'][:60]}...")
        print(f"       템플릿: {m['template_question'][:60]}...")
        print()

    # 매핑 실패 - 설문 문항
    if mapping_results['unmatched_survey']:
        print(f"\n[매핑 실패 - 설문] {len(mapping_results['unmatched_survey'])}개 문항")
        print("-" * 70)
        for u in mapping_results['unmatched_survey']:
            print(f"  평균: {u['mean']:.2f} | {u['survey_question'][:60]}...")

    # 매핑 실패 - 템플릿 행
    if mapping_results['unmatched_template']:
        print(f"\n[매핑 실패 - 템플릿] {len(mapping_results['unmatched_template'])}개 행")
        print("-" * 70)
        print(f"  비어있는 J열: {mapping_results['unmatched_template']}")

    print("\n" + "=" * 70)
    print(f"결과 파일: {output_path}")
    print("=" * 70)

    # 검토 리포트 파일로도 저장
    report_path = output_path.parent / (output_path.stem + '_검토리포트.json')
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(mapping_results, f, ensure_ascii=False, indent=2)
    print(f"검토 리포트: {report_path}")


def process_all_results(work_folder: str):
    """작업 폴더의 results 폴더 결과를 템플릿에 입력합니다.

    Args:
        work_folder: 작업 폴더명 (예: '2024_상반기_신입사원_입문과정')
    """
    base_dir = Path(__file__).parent.parent.parent
    templates_dir = base_dir / 'templates'

    work_dir = base_dir / 'data' / work_folder

    if not work_dir.exists():
        print(f"작업 폴더가 존재하지 않습니다: {work_folder}")
        return

    # 템플릿 파일 찾기
    template_files = list(templates_dir.glob('*.xlsx'))
    if not template_files:
        print("템플릿 파일이 없습니다.")
        print(f"  템플릿 폴더: {templates_dir}")
        return

    template_path = template_files[0]  # 첫 번째 템플릿 사용
    print(f"템플릿: {template_path.name}\n")

    results_dir = work_dir / 'results'
    output_dir = work_dir / 'output'
    result_files = list(results_dir.glob('*_stats.json'))

    if not result_files:
        print(f"[{work_folder}] 처리할 결과 파일이 없습니다.")
        return

    print(f"[{work_folder}] {len(result_files)}개의 결과 파일을 처리합니다.")

    for result_file in result_files:
        print(f"\n처리 중: {result_file.name}")

        # JSON 결과 로드
        with open(result_file, 'r', encoding='utf-8') as f:
            results = json.load(f)

        # 출력 파일명 생성
        output_name = result_file.stem.replace('_stats', '') + '_결과.xlsx'
        output_path = output_dir / output_name

        # 템플릿에 결과 입력
        mapping_results = fill_template(template_path, results, output_path)

        # 검토 리포트 출력
        print_verification_report(mapping_results, output_path)


if __name__ == '__main__':
    process_all_results()
